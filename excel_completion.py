from pathlib import Path
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
INPUT_XLSX = BASE_DIR / "excel" / "registro_curso_amor_sexualidad2.xlsx"
OUTPUT_XLSX = BASE_DIR / "excel" / "registro_curso_amor_sexualidad2_rellenado.xlsx"
WARN_LOG = BASE_DIR / "excel_completion_warnings.txt"

SHEET_NAME = None  # None = hoja activa; o pon el nombre exacto, p.ej. "Hoja1"

COL_NOMBRE = "Nombre"
COL_APELLIDOS = "Apellidos"
COL_CORREO = "Correo"
COL_USUARIO = "Usuario"
COL_CONTRASENA = "Contraseña"

def is_blank(value) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")

def email_local_part(email: str) -> str:
    email = (email or "").strip()
    if "@" not in email:
        raise ValueError(f"Email inválido (sin @): {email!r}")
    return email.split("@", 1)[0].strip()

def first_name(nombre: str) -> str:
    nombre = (nombre or "").strip()
    if not nombre:
        raise ValueError("Nombre vacío, no puedo generar contraseña.")
    return nombre.split()[0]

def main():
    wb = load_workbook(INPUT_XLSX)
    ws = wb[SHEET_NAME] if SHEET_NAME else wb.active

    # Detectar cabeceras (normalizando espacios)
    headers = {}
    header_row = 1
    for col_idx in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col_idx).value
        if isinstance(v, str) and v.strip():
            headers[v.strip()] = col_idx

    for required in (COL_NOMBRE, COL_APELLIDOS, COL_CORREO, COL_USUARIO, COL_CONTRASENA):
        if required not in headers:
            raise KeyError(f"No encuentro la columna {required!r}. Cabeceras detectadas: {list(headers.keys())}")

    c_nombre = headers[COL_NOMBRE]
    c_apellidos = headers[COL_APELLIDOS]
    c_correo = headers[COL_CORREO]
    c_usuario = headers[COL_USUARIO]
    c_contra = headers[COL_CONTRASENA]

    # 1) Validar unicidad de email (avisar pero no abortar)
    seen = {}
    duplicates = {}
    for r in range(2, ws.max_row + 1):
        correo = ws.cell(row=r, column=c_correo).value
        if is_blank(correo):
            continue
        correo_norm = str(correo).strip().lower()
        if correo_norm in seen:
            duplicates.setdefault(correo_norm, [seen[correo_norm]]).append(r)
        else:
            seen[correo_norm] = r

    warnings = []
    if duplicates:
        warnings.append("Se encontraron emails duplicados (se usará solo la primera aparición):")
        for email, rows in duplicates.items():
            warnings.append(f" - {email} en filas {rows}")
        
        # 2) Revisar si los duplicados tienen nombres/apellidos diferentes
        warnings.append("\n⚠ CHEQUEO DE DISCREPANCIAS EN DUPLICADOS:")
        for email, rows in duplicates.items():
            nombres = []
            for r in rows:
                nombre = ws.cell(row=r, column=c_nombre).value
                apellidos = ws.cell(row=r, column=c_apellidos).value
                nombres.append({
                    'fila': r,
                    'nombre': (nombre or "").strip(),
                    'apellidos': (apellidos or "").strip(),
                    'nombre_completo': f"{(nombre or '').strip()} {(apellidos or '').strip()}".strip()
                })
            
            # Comparar si son idénticos
            nombre_ref = nombres[0]['nombre_completo']
            discrepancia = False
            for n in nombres[1:]:
                if n['nombre_completo'].lower() != nombre_ref.lower():
                    discrepancia = True
                    break
            
            if discrepancia:
                warnings.append(f"  ⚠ {email}:")
                for n in nombres:
                    warnings.append(f"      Fila {n['fila']}: {n['nombre_completo']}")
                warnings.append(f"      → REVISAR: ¿Mismo usuario o dos personas distintas?")

        txt = "\n".join(warnings)
        print(txt)
        try:
            WARN_LOG.write_text(txt + "\n", encoding="utf-8")
        except Exception:
            pass

    # 2) Rellenar Usuario y Contraseña cuando falte Usuario
    changed = 0
    for r in range(2, ws.max_row + 1):
        usuario_val = ws.cell(row=r, column=c_usuario).value
        if not is_blank(usuario_val):
            continue

        correo = ws.cell(row=r, column=c_correo).value
        nombre = ws.cell(row=r, column=c_nombre).value

        if is_blank(correo):
            continue

        u = email_local_part(str(correo))
        ws.cell(row=r, column=c_usuario).value = u

        if not is_blank(nombre):
            ws.cell(row=r, column=c_contra).value = f"{first_name(str(nombre))}+A1+-"

        changed += 1

    wb.save(OUTPUT_XLSX)
    msg_final = f"OK. Filas actualizadas: {changed}. Guardado en: {OUTPUT_XLSX}"
    if duplicates:
        msg_final += " [⚠ AVISO: Emails duplicados rellenados en todas sus filas; solo el primero se registrará en Moodle]"
    print(msg_final)

if __name__ == "__main__":
    main()
