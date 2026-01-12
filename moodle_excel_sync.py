#!/usr/bin/env python3
"""
Script simple para crear/editar usuarios en Moodle desde Excel
"""

from pathlib import Path
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time
from datetime import datetime
import os

try:
    from dotenv import load_dotenv
except Exception:  # pragma: no cover
    load_dotenv = None

BASE_DIR = Path(__file__).resolve().parent
EXCEL_FILE = BASE_DIR / 'excel' / 'registro_curso_amor_sexualidad4_pendientes_moodle.xlsx'
RUN_TS = datetime.now().strftime("%Y%m%d_%H%M%S")
LOG_DIR = BASE_DIR / "logs"
LOG_DIR.mkdir(exist_ok=True)
LOG_FILE = LOG_DIR / f"log_moodle_sync__{EXCEL_FILE.stem}__{RUN_TS}.txt"

if load_dotenv is not None:
    load_dotenv(BASE_DIR / ".env")

MOODLE_BASE_URL = os.getenv("MOODLE_BASE_URL", "https://campus.edufamilia.com").rstrip("/")
MOODLE_ADMIN_USER = os.getenv("MOODLE_ADMIN_USER", "")
MOODLE_ADMIN_PASSWORD = os.getenv("MOODLE_ADMIN_PASSWORD", "")

# ===== CONFIGURACIÓN DE FILAS A PROCESAR =====
# Define qué filas del Excel procesará el script
# Si FILAS_A_PROCESAR es None, se procesará desde FILA_INICIO hasta la última fila del Excel
FILA_INICIO = 2  # Encabezados en fila 1
FILAS_A_PROCESAR = None  # Procesa todas las filas desde FILA_INICIO hasta el final
# ============================================

# Si un nombre/apellido viene TODO EN MAYÚSCULAS, Moodle no debería fallar por eso,
# pero a veces es preferible normalizarlo para evitar resultados feos en la UI.
# Esta normalización SOLO se aplica cuando el texto parece estar en mayúsculas.
NORMALIZAR_MAYUSCULAS_A_TITULO = True


def _solo_mayusculas(texto: str) -> bool:
    letras = [c for c in texto if c.isalpha()]
    return bool(letras) and all(c.isupper() for c in letras)


def _normalizar_nombre(texto: str) -> str:
    s = (texto or "").strip()
    if not NORMALIZAR_MAYUSCULAS_A_TITULO:
        return s
    if _solo_mayusculas(s):
        return s.title()
    return s


def _extraer_errores_moodle(driver) -> list[str]:
    """Intenta capturar mensajes de error visibles tras guardar un formulario."""
    errores: list[str] = []
    try:
        candidates = driver.find_elements(
            By.XPATH,
            "//div[contains(@class,'alert-danger') or contains(@class,'error') or @role='alert']"
            " | //span[contains(@class,'error')]"
            " | //div[contains(@class,'invalid-feedback')]",
        )
        for el in candidates:
            try:
                t = (el.text or "").strip()
            except Exception:
                t = ""
            if t and t not in errores:
                errores.append(t)
    except Exception:
        return []
    return errores


def _buscar_email_en_listado(driver, email: str) -> bool:
    """Busca un email en la tabla de usuarios (admin/user.php) tras filtrar por email."""
    wait = WebDriverWait(driver, 15)
    driver.get(f"{MOODLE_BASE_URL}/admin/user.php")
    time.sleep(2)

    try:
        mostrar_mas = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a.moreless-toggler")))
        mostrar_mas.click()
        time.sleep(1)
    except Exception:
        pass

    campo_email = wait.until(EC.presence_of_element_located((By.ID, "id_email")))
    campo_email.clear()
    campo_email.send_keys(email)
    campo_email.send_keys(Keys.RETURN)
    time.sleep(2)

    # 1) Caso "No se encuentran usuarios"
    if driver.find_elements(By.XPATH, "//*[contains(text(), 'No se encuentran usuarios')]"):
        return False

    # 2) Caso tabla con el email presente
    if driver.find_elements(By.XPATH, f"//*[contains(normalize-space(.), '{email}')]"):
        return True

    return False

def log_msg(mensaje):
    """Imprime mensaje con timestamp tanto en consola como en log"""
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    linea = f"[{ts}] {mensaje}"
    print(mensaje)
    with open(LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(linea + '\n')

def leer_registros_excel(filas):
    """Lee los datos de los registros especificados del Excel"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    registros = []
    for fila in filas:
        apellidos = ws.cell(fila, 1).value  # Columna 1: Apellidos
        nombre = ws.cell(fila, 2).value     # Columna 2: Nombre
        email = ws.cell(fila, 3).value      # Columna 3: Email
        usuario = ws.cell(fila, 6).value    # Columna 6: Usuario
        contrasena = ws.cell(fila, 7).value # Columna 7: Contraseña
        
        if nombre and apellidos and email and usuario:
            registros.append({
                'fila': fila,
                'nombre': _normalizar_nombre(nombre),
                'apellidos': _normalizar_nombre(apellidos),
                'email': email.strip(),
                'usuario': usuario.strip().lower(),
                'contrasena': contrasena.strip() if contrasena else None
            })
    
    return registros

def obtener_filas_desde(inicio: int):
    """Devuelve lista de filas desde 'inicio' hasta el final del Excel"""
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
    ws = wb.active
    max_row = ws.max_row
    wb.close()
    filas = list(range(inicio, max_row + 1))
    log_msg(f"Rango detectado: fila {inicio} hasta fila {max_row} (total: {len(filas)} filas)")
    return filas

def login_moodle(driver):
    """Inicia sesión en Moodle"""
    log_msg("\n✓ Iniciando sesión en Moodle...")

    if not MOODLE_ADMIN_USER or not MOODLE_ADMIN_PASSWORD:
        raise RuntimeError(
            "Faltan credenciales. Define MOODLE_ADMIN_USER y MOODLE_ADMIN_PASSWORD en el entorno o en el archivo .env"
        )

    driver.get(f"{MOODLE_BASE_URL}/admin/search.php")
    time.sleep(2)
    
    wait = WebDriverWait(driver, 15)
    
    # Usuario
    campo_usuario = wait.until(EC.presence_of_element_located((By.ID, "username")))
    campo_usuario.send_keys(MOODLE_ADMIN_USER)
    
    # Contraseña
    campo_password = driver.find_element(By.ID, "password")
    campo_password.send_keys(MOODLE_ADMIN_PASSWORD)
    
    # Click en Log in
    login_btn = driver.find_element(By.XPATH, "//button[contains(text(), 'Log in')]")
    login_btn.click()
    
    time.sleep(4)
    log_msg("✓ Sesión iniciada")

def procesar_usuario(driver, registro, es_primero=False):
    """Verifica si el usuario existe por email y lo crea o edita según corresponda"""
    fila = registro['fila']
    nombre = registro['nombre']
    apellidos = registro['apellidos']
    email = registro['email']
    usuario = registro['usuario']
    contrasena = registro['contrasena']
    
    log_msg(f"\n[Fila {fila}] Procesando: {nombre} {apellidos} ({email})")
    
    wait = WebDriverWait(driver, 15)
    
    try:
        # 1. Ir a "Examinar lista de usuarios"
        driver.get(f"{MOODLE_BASE_URL}/admin/user.php")
        time.sleep(2)
        
        # 2. Si NO es el primero, eliminar filtros anteriores
        if not es_primero:
            try:
                eliminar_filtros = wait.until(EC.element_to_be_clickable((By.ID, "id_removeall")))
                eliminar_filtros.click()
                time.sleep(1)
            except:
                pass  # Si no hay filtros, continuar
        
        # 3. Hacer clic en "Mostrar más..."
        mostrar_mas = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a.moreless-toggler")))
        mostrar_mas.click()
        time.sleep(1)
        
        # 4. Buscar por email
        campo_email = wait.until(EC.presence_of_element_located((By.ID, "id_email")))
        campo_email.clear()
        campo_email.send_keys(email)
        campo_email.send_keys(Keys.RETURN)
        time.sleep(3)
        
        # 5. Verificar si encuentra usuarios (sin usar except genérico)
        if driver.find_elements(By.XPATH, "//*[contains(text(), 'No se encuentran usuarios')]"):
            log_msg(f"  → Usuario NO existe. Creando...")

            boton_crear = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Crear un nuevo usuario')]"))
            )
            boton_crear.click()
            time.sleep(2)

            ok = crear_usuario_en_formulario(driver, registro)
            if not ok:
                return "error"

            # Verificación post-guardado
            if _buscar_email_en_listado(driver, email):
                log_msg("  ✓ Verificación: email aparece en la lista")
                return "created"
            log_msg("  ✗ Verificación fallida: no aparece el email en la lista")
            return "error"

        # Si no hay mensaje de no-encontrado, intentamos editar.
        log_msg(f"  → Usuario YA existe (o no se mostró el mensaje). Editando...")

        edit_icon = wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href*='user/editadvanced.php'] i.fa-cog"))
        )
        edit_icon.click()
        time.sleep(2)

        campo_nombre = wait.until(EC.presence_of_element_located((By.ID, "id_firstname")))
        campo_nombre.clear()
        campo_nombre.send_keys(nombre)

        campo_apellido = driver.find_element(By.ID, "id_lastname")
        campo_apellido.clear()
        campo_apellido.send_keys(apellidos)

        boton_guardar = driver.find_element(By.ID, "id_submitbutton")
        driver.execute_script("arguments[0].scrollIntoView(true);", boton_guardar)
        time.sleep(1)
        boton_guardar.click()
        time.sleep(3)

        errores = _extraer_errores_moodle(driver)
        if errores:
            for err in errores[:3]:
                log_msg(f"  ✗ Error Moodle: {err[:160]}")
            return "error"

        if _buscar_email_en_listado(driver, email):
            log_msg("  ✓ Verificación: email aparece en la lista")
            log_msg(f"  ✓✓ Usuario editado exitosamente")
            return "edited"

        log_msg("  ⚠ Editado sin poder verificar en lista (posible filtro/export)")
        return "edited"
            
    except Exception as e:
        msg = str(e)
        log_msg(f"  ✗ Error: {msg[:120]}")
        return "error"

def crear_usuario_en_formulario(driver, registro):
    """Crea un usuario cuando ya estamos en el formulario de creación"""
    try:
        wait = WebDriverWait(driver, 15)
        
        # 1. Esperar y rellenar Nombre de usuario
        campo_usuario = wait.until(EC.presence_of_element_located((By.ID, "id_username")))
        campo_usuario.clear()
        campo_usuario.send_keys(registro['usuario'])
        time.sleep(0.5)
        log_msg(f"  Nombre de usuario: {registro['usuario']}")
        
        # 2. Rellenar Dirección de correo
        campo_email = driver.find_element(By.ID, "id_email")
        campo_email.clear()
        campo_email.send_keys(registro['email'])
        time.sleep(0.5)
        log_msg(f"  Email: {registro['email']}")
        
        # 3. Rellenar Nombre
        campo_nombre = driver.find_element(By.ID, "id_firstname")
        campo_nombre.clear()
        campo_nombre.send_keys(registro['nombre'])
        time.sleep(0.5)
        log_msg(f"  Nombre: {registro['nombre']}")
        
        # 4. Rellenar Apellidos
        campo_apellido = driver.find_element(By.ID, "id_lastname")
        campo_apellido.clear()
        campo_apellido.send_keys(registro['apellidos'])
        time.sleep(0.5)
        log_msg(f"  Apellidos: {registro['apellidos']}")
        
        # 5. Hacer clic en "Haz click para insertar texto" para la contraseña
        log_msg(f"  ⏳ Buscando enlace de contraseña...")
        try:
            enlace_contrasena = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[@data-passwordunmask='edit']")))
            driver.execute_script("arguments[0].scrollIntoView(true);", enlace_contrasena)
            time.sleep(1)
            driver.execute_script("arguments[0].click();", enlace_contrasena)
            log_msg(f"  ✓ Click en enlace de contraseña")
            time.sleep(2)
        except Exception as e:
            log_msg(f"  ⚠ Error con enlace de contraseña: {str(e)[:50]}")
            pass
        
        # 6. Rellenar Nueva contraseña
        try:
            campo_password = wait.until(EC.presence_of_element_located((By.ID, "id_newpassword")))
            driver.execute_script("arguments[0].scrollIntoView(true);", campo_password)
            time.sleep(1)
            campo_password.clear()
            campo_password.send_keys(registro['contrasena'])
            time.sleep(0.5)
            log_msg(f"  Contraseña: {registro['contrasena']}")
        except Exception as e:
            log_msg(f"  ✗ Error al rellenar contraseña: {str(e)[:50]}")
            raise
        
        # 7. Hacer clic en "Crear usuario"
        submit_btn = wait.until(EC.element_to_be_clickable((By.NAME, "submitbutton")))
        driver.execute_script("arguments[0].scrollIntoView(true);", submit_btn)
        time.sleep(1)
        submit_btn.click()
        log_msg(f"  ✓ Haciendo click en 'Crear Usuario'")
        time.sleep(3)
        
        # 8. Verificar errores del formulario
        errores = _extraer_errores_moodle(driver)
        if errores:
            for err in errores[:3]:
                log_msg(f"  ✗ Error Moodle: {err[:160]}")
            return False

        log_msg(f"  ✓✓ Usuario creado exitosamente")
        return True
            
    except Exception as e:
        log_msg(f"  ✗ Error: {str(e)[:100]}")
        return False

def crear_usuario_moodle(driver, registro):
    """Crea un usuario en Moodle"""
    fila = registro['fila']
    nombre = registro['nombre']
    apellidos = registro['apellidos']
    email = registro['email']
    usuario = registro['usuario']
    contrasena = registro['contrasena']
    
    log_msg(f"\n[Fila {fila}] Creando usuario: {usuario}")
    log_msg(f"  Nombre: {nombre} | Apellidos: {apellidos} | Email: {email}")
    
    crear_usuario_url = f"{MOODLE_BASE_URL}/user/editadvanced.php?id=-1"
    driver.get(crear_usuario_url)
    time.sleep(2)
    
    wait = WebDriverWait(driver, 15)
    
    try:
        # Campo usuario
        campo_usuario = wait.until(EC.presence_of_element_located((By.ID, "id_username")))
        campo_usuario.clear()
        campo_usuario.send_keys(usuario)
        
        # Campo email
        campo_email = driver.find_element(By.ID, "id_email")
        campo_email.clear()
        campo_email.send_keys(email)
        
        # Campo nombre
        campo_nombre = driver.find_element(By.ID, "id_firstname")
        campo_nombre.clear()
        campo_nombre.send_keys(nombre)
        
        # Campo apellido
        campo_apellido = driver.find_element(By.ID, "id_lastname")
        campo_apellido.clear()
        campo_apellido.send_keys(apellidos)
        
        # Habilitar contraseña
        try:
            enlace_pwd = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//a[@data-passwordunmask='edit']"))
            )
            driver.execute_script("arguments[0].scrollIntoView(true);", enlace_pwd)
            time.sleep(1)
            driver.execute_script("arguments[0].click();", enlace_pwd)
            time.sleep(2)
        except:
            pass
        
        # Llenar contraseña
        try:
            campo_pwd = wait.until(EC.presence_of_element_located((By.ID, "id_newpassword")))
            time.sleep(1)
            driver.execute_script("arguments[0].scrollIntoView(true);", campo_pwd)
            campo_pwd.send_keys(contrasena)
        except:
            log_msg(f"  ⚠ Error en contraseña")
        
        # Click en Crear Usuario
        boton_crear = wait.until(EC.element_to_be_clickable((By.NAME, "submitbutton")))
        boton_crear.click()
        time.sleep(3)
        
        errores = _extraer_errores_moodle(driver)
        if errores:
            for err in errores[:3]:
                log_msg(f"  ✗ Error Moodle: {err[:160]}")
            return False

        if _buscar_email_en_listado(driver, email):
            log_msg("  ✓ Verificación: email aparece en la lista")
            log_msg(f"  ✓✓ Usuario creado exitosamente")
            return True

        log_msg("  ✗ Verificación fallida: no aparece el email en la lista")
        return False
            
    except Exception as e:
        log_msg(f"  ✗ Error: {str(e)[:100]}")
        return False

def editar_usuario_moodle(driver, fila, nombre, apellidos, email):
    """Edita un usuario existente en Moodle buscando por email"""
    log_msg(f"\n[Fila {fila}] Editando usuario...")
    log_msg(f"  Nombre: {nombre} | Apellidos: {apellidos} | Email: {email}")
    
    wait = WebDriverWait(driver, 15)
    
    try:
        # 1. Ir a "Examinar lista de usuarios"
        driver.get(f"{MOODLE_BASE_URL}/admin/user.php")
        time.sleep(2)
        
        # 2. Hacer clic en "Mostrar más..."
        mostrar_mas = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a.moreless-toggler")))
        mostrar_mas.click()
        time.sleep(1)
        
        # 3. Escribir email en el campo "Dirección de correo"
        campo_email = wait.until(EC.presence_of_element_located((By.ID, "id_email")))
        campo_email.clear()
        campo_email.send_keys(email)
        campo_email.send_keys(Keys.RETURN)
        time.sleep(3)
        
        # 4. Hacer clic en el icono de configuración (engranaje) para editar
        edit_icon = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href*='user/editadvanced.php'] i.fa-cog")))
        edit_icon.click()
        time.sleep(2)
        
        # 5. Modificar Nombre
        campo_nombre = wait.until(EC.presence_of_element_located((By.ID, "id_firstname")))
        campo_nombre.clear()
        campo_nombre.send_keys(nombre)
        
        # 6. Modificar Apellidos
        campo_apellido = driver.find_element(By.ID, "id_lastname")
        campo_apellido.clear()
        campo_apellido.send_keys(apellidos)
        
        # 7. Guardar cambios
        boton_guardar = driver.find_element(By.ID, "id_submitbutton")
        driver.execute_script("arguments[0].scrollIntoView(true);", boton_guardar)
        time.sleep(1)
        boton_guardar.click()
        time.sleep(3)
        
        log_msg(f"  ✓✓ Usuario editado exitosamente")
        return True
        
    except Exception as e:
        log_msg(f"  ✗ Error: {str(e)[:100]}")
        return False

def main():
    """Función principal"""
    log_msg("=" * 80)
    log_msg("PROCESAR USUARIOS EN MOODLE")
    log_msg("=" * 80)
    log_msg(f"Excel: {EXCEL_FILE.name}")
    log_msg(f"Log: {LOG_FILE.name}")
    
    # Configurar Chrome
    chrome_options = Options()
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_argument("user-agent=Mozilla/5.0")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
    
    try:
        # Login
        login_moodle(driver)
        
        # Procesar registros según configuración
        filas_a_procesar = FILAS_A_PROCESAR or obtener_filas_desde(FILA_INICIO)
        registros = leer_registros_excel(filas_a_procesar)
        
        log_msg(f"\nProcesando {len(registros)} registros...")
        for r in registros:
            log_msg(f"  - Fila {r['fila']}: {r['nombre']} {r['apellidos']}")
        
        created = 0
        edited = 0
        errors = 0
        for idx, registro in enumerate(registros):
            es_primero = (idx == 0)
            result = procesar_usuario(driver, registro, es_primero)
            if result == "created":
                created += 1
            elif result == "edited":
                edited += 1
            else:
                errors += 1
            time.sleep(1)
        
        log_msg("\n" + "=" * 80)
        log_msg("✓ Proceso completado")
        log_msg(f"Resumen: creados={created}, editados={edited}, errores={errors}, total={len(registros)}")
        log_msg("=" * 80)
        
    except Exception as e:
        log_msg(f"\n✗ Error general: {e}")
    finally:
        driver.quit()

if __name__ == "__main__":
    main()
